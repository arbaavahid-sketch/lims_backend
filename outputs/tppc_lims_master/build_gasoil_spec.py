"""پرکردن مشخصات (min/max) محصول «نفت گاز» به‌عنوان الگو + ساخت فایل import SENAITE.

حدود از استاندارد گازوئیل (EN 590 / INSO) — همه با برچسب «تأیید مدیر فنی لازم».
خروجی: TPPC_Specs_Gasoil.xlsx  (شیت Review برای بازبینی + شیت Analysis Specifications برای import)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SERV = "TPPC_AnalysisServices_final.xlsx"
OUT = "TPPC_Specs_Gasoil.xlsx"
PRODUCT = "نفت گاز"

# حدود بر اساس keyword: (min, max, unit, standard, note)
# None = بدون حد در آن سمت. "" در min/max یعنی TBD (فصلی/نیازمند داده).
LIMITS = {
    "AS_49176_065": (None, 0.01, "% mass", "ASTM D482 / EN 590", ""),                 # خاکستر
    "AS_30884_120": (None, 2.0, "-", "ASTM D1500 / INSO 4903", "رنگ ASTM؛ گرید را تأیید کنید"),  # رنگ
    "AS_96932_069": (55, 999999, "°C", "ASTM D93 / EN 590", "بدون حد بالا"),             # نقطه اشتعال
    "AS_78253_021": (None, 1, "Class", "ISIRI 336 / ASTM D130", "کلاس ۱ (حداکثر)"),     # خوردگی نوار مسی
    "AS_55958_156": (2.0, 4.5, "mm²/s", "ASTM D445 / EN 590", ""),                      # گرانروی 40°C
    "AS_23337_008": (None, 200, "mg/kg", "ASTM D6304 / EN 590", ""),                    # آب کارل فیشر
    "AS_45447_080": (820, 845, "kg/m³", "ASTM D1298 / EN 590", ""),                     # چگالی 15°C
    "AS_27082_143": (46, 999999, "-", "INSO 8525 / EN 590", "بدون حد بالا"),             # شاخص ستان
    "AS_53330_126": ("", "", "°C", "ASTM D2500 / INSO 5438", "فصلی — حد را بر اساس فصل/گرید وارد کنید"),  # نقطه ابری
    "AS_99532_110": (None, 1, "Class", "ISIRI 336 / ASTM D130", "کلاس ۱ (حداکثر)"),     # خوردگی تیغه مسی
    "AS_24745_103": ("", "", "°C", "ASTM D97 / INSO 201", "فصلی — حد را بر اساس فصل/گرید وارد کنید"),   # نقطه ریزش
    "AS_79390_068": (55, 999999, "°C", "ASTM D93 / EN 590", "تکراری؛ بدون حد بالا"),     # نقطه اشتعال (تکراری)
}

# عنوان دقیق سرویس‌ها را از سرویس‌های واردشده بگیر (برای تطبیق import)
wb = openpyxl.load_workbook(SERV, data_only=True)
ws = wb["Analysis Services"]
rows = list(ws.iter_rows(values_only=True))
hi = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Analysis Keyword")
hdr = [str(h).strip() if h else "" for h in rows[hi]]
svc = {}
for r in rows[hi+1:]:
    d = dict(zip(hdr, r))
    kw = d.get("Analysis Keyword")
    if kw:
        svc[str(kw).strip()] = str(d.get("Analysis Title FA") or "").strip()

out = openpyxl.Workbook()
hdr_fill = PatternFill("solid", fgColor="1A3C6E"); hdr_font = Font(color="FFFFFF", bold=True)
note_fill = PatternFill("solid", fgColor="FFF6E5")

# --- شیت Review (برای بازبینی مدیر فنی) ---
rv = out.active; rv.title = "Review"
cols = ["آزمون", "Keyword", "حداقل", "حداکثر", "واحد", "استاندارد", "یادداشت"]
for c, n in enumerate(cols, 1):
    cell = rv.cell(1, c, n); cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center")
for ri, (kw, (mn, mx, unit, std, note)) in enumerate(LIMITS.items(), 2):
    title = svc.get(kw, "(سرویس یافت نشد)")
    vals = [title, kw,
            "" if mn is None else mn, "" if mx is None else mx,
            unit, std, note]
    for ci, v in enumerate(vals, 1):
        cell = rv.cell(ri, ci, v)
        if note:
            cell.fill = note_fill
for i, w in enumerate([40, 16, 10, 10, 10, 26, 40], 1):
    rv.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
rv.freeze_panes = "A2"

# --- شیت Analysis Specifications (فرمت import SENAITE، هدر ردیف ۱، داده ردیف ۴) ---
sp = out.create_sheet("Analysis Specifications")
keys = ["Title", "Client_title", "SampleType_title", "service", "min", "max", "error"]
for c, k in enumerate(keys, 1):
    cell = sp.cell(1, c, k); cell.fill = hdr_fill; cell.font = hdr_font
sp.cell(2, 1, "Analysis Specifications")
for c, k in enumerate(keys, 1):
    sp.cell(3, c, k)
ri = 4
imported = skipped = 0
for kw, (mn, mx, unit, std, note) in LIMITS.items():
    title = svc.get(kw)
    if not title or (mn in ("", None) and mx in ("", None)):
        skipped += 1
        continue  # سرویس نامشخص یا حد TBD → در import نمی‌آید
    sp.cell(ri, 1, PRODUCT)            # Title == نام نوع نمونه → spec پیش‌فرض همان محصول
    sp.cell(ri, 3, PRODUCT)           # SampleType_title
    sp.cell(ri, 4, title)             # service (عنوان سرویس)
    sp.cell(ri, 5, "" if mn is None else mn)
    sp.cell(ri, 6, "" if mx is None else mx)
    sp.cell(ri, 7, "")                # error/tolerance
    ri += 1
    imported += 1
sp.freeze_panes = "A2"

out.save(OUT)
print(f"✅ {OUT} ساخته شد")
print(f"   Review: 12 آزمون | Import: {imported} آزمون آماده، {skipped} کنار گذاشته‌شده (TBD/فصلی)")
print("\n=== جدول بازبینی ===")
for kw, (mn, mx, unit, std, note) in LIMITS.items():
    title = svc.get(kw, "؟")[:36]
    rng = f"{'' if mn in (None,'') else mn} .. {'' if mx in (None,'') else mx}".strip()
    flag = "  ⚠️ "+note if note else ""
    print(f"  {title:36} | {rng or 'TBD':12} {unit:7} | {std}{flag}")
