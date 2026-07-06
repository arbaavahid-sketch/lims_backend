"""Build reviewed sample-type workbook for Phase 1 product tree cleanup.

Input:
    TPPC_LIMS_Master_Data_units_filled.xlsx / Sample Types

Output:
    TPPC_SampleTypes_reviewed.xlsx

The output intentionally separates review decisions from import-ready sample
types. Broad source-scope labels are kept for traceability but are not included
in the import sheet unless the lab confirms them as real incoming sample names.
"""
from collections import Counter
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


SRC = "TPPC_LIMS_Master_Data_units_filled.xlsx"
OUT = "TPPC_SampleTypes_reviewed.xlsx"


BROAD_REVIEW = {
    "فرآورده های نفتی": "Broad family label; do not import as daily sample type without confirmation.",
    "فرآورده های میوه ها و سبزی ها": "Appears outside petroleum scope; likely source-data noise.",
    "سایر فرآورده های تقطیری": "Catch-all wording; map to a specific distillate family first.",
    "نفت خام و فرآورده های نفتی": "Combines family and product wording.",
    "فرآورده های نفتی و سوخت های مایع": "Overlaps with fuel/distillate families.",
    "موتور گازسوز": "Looks like equipment/application wording rather than sample material.",
    "پرایمر اصلاح شده": "Needs confirmation: bitumen/coating product or separate sample type.",
    "گاز": "Gas sampling may need separate workflow/container decision.",
    "نفت": "Too broad unless used as a real request sample type.",
    "هیدروکربن": "Too broad unless used as a real request sample type.",
    "هیدروکربن ها": "Too broad unless used as a real request sample type.",
    "حلال": "Generic family label; prefer specific solvent sample types.",
    "روغن موتور": "Generic family label; prefer gasoline/diesel/grade-level oil types.",
    "روغن روانکار": "Generic family label; prefer application/grade-level oil types.",
    "روغن های روان کننده": "Generic family label; prefer application/grade-level oil types.",
    "روان کارها": "Generic family label; prefer application/grade-level oil types.",
    "فرآورده های نفتی و حلال های هیدروکربنی": "Mixed broad label; split into product families.",
}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def prefix_from_code(code, index):
    value = re.sub(r"[^A-Za-z0-9]", "", clean(code)).upper()
    if value:
        return value[:8]
    return "ST{:03d}".format(index)


def classify(title):
    title = clean(title)
    if any(k in title for k in ("خنک", "گلایکول", "گلیکول", "ضدیخ")):
        return "ضدیخ / مایعات خنک کننده"
    if any(k in title for k in ("سوخت", "بنزین", "نفتا", "نفت سفید", "بیودیزل", "جت", "دیزل", "توربین")):
        return "سوخت، نفتا و فرآورده های تقطیری"
    if any(k in title for k in ("روغن", "روان کار", "روانکار", "گریس")):
        return "روغن و روانکار"
    if any(k in title for k in ("حلال", "هیدروکربن", "پارافین")):
        return "حلال و هیدروکربن"
    if "قیر" in title:
        return "قیر و مواد قیری"
    if any(k in title for k in ("نفت خام", "نفت کوره", "نفت گاز", "نفت و گاز")):
        return "نفت خام و محصولات سنگین"
    if any(k in title for k in ("فرآورده های نفتی", "مایعات نفتی", "نفت")):
        return "عمومی / بالادستی"
    return "سایر / نیازمند تصمیم"


def decision_for(row):
    title = clean(row.get("Sample Type / Product"))
    status = clean(row.get("Status")).lower()
    if status != "active":
        return "exclude", "", "Inactive in source master data."
    if title in BROAD_REVIEW:
        return "review", "", BROAD_REVIEW[title]
    return "import", title, "Specific enough for selectable sample type."


def style_header(ws):
    fill = PatternFill("solid", fgColor="1A3C6E")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Sample Types"]
    rows = list(ws.iter_rows(values_only=True))
    header_index = next(i for i, row in enumerate(rows) if row and clean(row[0]) == "Sample Type Code")
    headers = [clean(value) for value in rows[header_index]]
    source_rows = [
        dict(zip(headers, row))
        for row in rows[header_index + 1:]
        if any(value not in (None, "") for value in row)
    ]

    reviewed = []
    used_prefixes = set()
    for index, row in enumerate(source_rows, 1):
        family = classify(row.get("Sample Type / Product"))
        decision, import_title, reason = decision_for(row)
        prefix = prefix_from_code(row.get("Sample Type Code"), index)
        while prefix in used_prefixes:
            prefix = "{}{}".format(prefix[:6], index)
        used_prefixes.add(prefix)
        reviewed.append({
            "Source Code": clean(row.get("Sample Type Code")),
            "Source Title": clean(row.get("Sample Type / Product")),
            "Family": family,
            "Source Row Count": row.get("Source Row Count") or 0,
            "Source Status": clean(row.get("Status")),
            "Decision": decision,
            "Import Title": import_title,
            "Import Prefix": prefix,
            "Reason": reason,
        })

    out = openpyxl.Workbook()
    summary = out.active
    summary.title = "Summary"
    summary_rows = [
        ("Metric", "Value"),
        ("Source sample/product rows", len(reviewed)),
        ("Import-ready sample types", sum(1 for row in reviewed if row["Decision"] == "import")),
        ("Review before import", sum(1 for row in reviewed if row["Decision"] == "review")),
        ("Excluded inactive", sum(1 for row in reviewed if row["Decision"] == "exclude")),
    ]
    family_counts = Counter(row["Family"] for row in reviewed)
    for r, values in enumerate(summary_rows, 1):
        summary.append(values)
    summary.append(("", ""))
    summary.append(("Family", "Rows"))
    for family, count in family_counts.most_common():
        summary.append((family, count))
    style_header(summary)
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 18

    review = out.create_sheet("Review Decisions")
    review_headers = [
        "Source Code", "Source Title", "Family", "Source Row Count", "Source Status",
        "Decision", "Import Title", "Import Prefix", "Reason",
    ]
    review.append(review_headers)
    for row in reviewed:
        review.append([row[key] for key in review_headers])
    style_header(review)
    for col, width in {
        "A": 16, "B": 44, "C": 34, "D": 14, "E": 14,
        "F": 12, "G": 44, "H": 14, "I": 64,
    }.items():
        review.column_dimensions[col].width = width
    review.freeze_panes = "A2"

    import_ws = out.create_sheet("Sample Types Import")
    import_headers = [
        "title", "description", "RetentionPeriod", "Hazardous",
        "SampleMatrix_title", "Prefix", "MinimumVolume", "ContainerType_title",
    ]
    import_ws.append(import_headers)
    import_ws.append(["Sample Types", "", "", "", "", "", "", ""])
    import_ws.append(import_headers)
    for row in reviewed:
        if row["Decision"] != "import":
            continue
        import_ws.append([
            row["Import Title"],
            "Family: {}".format(row["Family"]),
            "",
            "0",
            "",
            row["Import Prefix"],
            "0 mL",
            "",
        ])
    style_header(import_ws)
    for col, width in {
        "A": 44, "B": 36, "C": 18, "D": 10,
        "E": 20, "F": 14, "G": 16, "H": 22,
    }.items():
        import_ws.column_dimensions[col].width = width
    import_ws.freeze_panes = "A4"

    out.save(OUT)
    counts = Counter(row["Decision"] for row in reviewed)
    print("OK created: {}".format(OUT))
    print("source rows: {}".format(len(reviewed)))
    print("import: {} | review: {} | exclude: {}".format(
        counts.get("import", 0),
        counts.get("review", 0),
        counts.get("exclude", 0),
    ))


if __name__ == "__main__":
    main()
