"""نهایی‌سازی و پاک‌سازی شیت Instruments برای ورود به SENAITE.

ورودی : TPPC_LIMS_Master_Data_units_filled.xlsx  (شیت Instruments)
خروجی: TPPC_Instruments_final.xlsx  (شیت‌های Instruments + QA_Notes)

کارها: جداسازی/اصلاح سازنده و کشور، اصلاح املای انگلیسی، افزودن نوع دستگاه،
محاسبهٔ سررسید کالیبراسیون بعدی. سریال/تاریخ/محل/مسئول دست‌نخورده می‌مانند.
مواردی که نیاز به دادهٔ واقعی کاربر دارند در شیت QA_Notes فهرست می‌شوند.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

SRC = "TPPC_LIMS_Master_Data_units_filled.xlsx"
OUT = "TPPC_Instruments_final.xlsx"

# اصلاح سازنده/کشور، نام انگلیسی، نوع دستگاه — کلید = Instrument Code
# (Manufacturer, Country_FA, NameEN_fixed, InstrumentType)
FIX = {
    "GC_DHA":            ("Chromatec", "روسیه", "Gas Chromatograph (DHA)", "کروماتوگراف گازی"),
    "ICP":              ("Agilent", "آمریکا", "ICP-OES Spectrometer", "طیف‌سنج ICP-OES"),
    "RVP":              ("TERMEX", "روسیه", "Reid Vapor Pressure Tester", "دستگاه فشار بخار رید"),
    "COND":             ("Mettler Toledo", "چین", "Conductivity Meter", "هدایت‌سنج"),
    "FP_01":            ("NXA", "روسیه", "Flash Point Tester", "دستگاه نقطه اشتعال"),
    "FP_02":            ("دانشور شیمی", "ایران", "Flash Point Tester", "دستگاه نقطه اشتعال"),
    "PH":               ("WTW", "آلمان", "pH Meter", "pH متر"),
    "VM":               ("TERMEX", "روسیه", "Kinematic Viscometer", "ویسکومتر"),
    "Turbidity":        ("Hach", "آلمان", "Turbidity Meter", "کدورت‌سنج"),
    "DR":               ("Hach", "آلمان", "Spectrophotometer", "اسپکتروفتومتر"),
    "CS":               ("TERMEX", "روسیه", "Copper Corrosion Bath", "حمام خوردگی مس"),
    "Distilation_01":   ("NXA", "روسیه", "Distillation Unit", "دستگاه تقطیر"),
    "Distilation_02":   ("دانشور شیمی", "ایران", "Distillation Unit", "دستگاه تقطیر"),
    "Pour_point":       ("TERMEX", "روسیه", "Pour Point Tester", "دستگاه نقطه ریزش"),
    "Cloud_point":      ("TERMEX", "روسیه", "Cloud Point Tester", "دستگاه نقطه ابری"),
    "Coulometry":       ("KEM", "ژاپن", "Karl Fischer Coulometer", "کولومتر کارل فیشر"),
    "Densitimeter":     ("TERMEX", "روسیه", "Density Meter", "چگالی‌سنج"),
    "SE":               ("Spectroscan", "روسیه", "EDXRF Sulfur Analyzer", "آنالایزر گوگرد XRF"),
    "OCTAN_CETAN_METER":("SHATOX", "روسیه", "Octane/Cetane Analyzer", "دستگاه عدد اکتان/ستان"),
    "PT":               ("KEM", "ژاپن", "Potentiometric Titrator", "تیتراتور پتانسیومتری"),
    "KARL_FISHER":      ("KEM", "ژاپن", "Karl Fischer Titrator", "تیتراتور کارل فیشر"),
    "TITRATOR":         ("Metrohm", "سوئیس", "Titrator", "تیتراتور"),
    "H2_Generator":     ("Chromatec", "روسیه", "Hydrogen Generator", "مولد هیدروژن"),
    "Catalitic_Purifier":("Chromatec", "روسیه", "Catalytic Purifier", "خالص‌ساز کاتالیزوری"),
    "INST_027":         ("Radwag", "لهستان", "Analytical Balance", "ترازوی آنالیتیک"),
    "heater_mantle":    ("Alfa", "ایران", "Heating Mantle", "هیتر مانتل"),
    "HEATER_STIRRER":   ("Alfa", "ایران", "Hot Plate Stirrer", "هیتر استیرر"),
    "Incubator":        ("Alfa", "ایران", "Incubator", "انکوباتور"),
    "saybolt":          ("Lovibond", "آلمان", "Saybolt Colorimeter", "دستگاه رنگ سیبولت"),
}

# مدل‌هایی که کاربر تأیید/اصلاح کرد
MODEL_OVERRIDE = {
    "Densitimeter": "VIP-2MP",
    "Distilation_01": "ARNS21",
    "FP_01": "ATV21",
}


def add_months(d, months):
    if not isinstance(d, datetime):
        return ""
    m = d.month - 1 + int(months)
    y = d.year + m // 12
    m = m % 12 + 1
    day = min(d.day, [31,29 if y%4==0 and (y%100!=0 or y%400==0) else 28,31,30,31,30,31,31,30,31,30,31][m-1])
    return datetime(y, m, day)


# دستگاه‌هایی که در فایل منبع نبودند و کاربر خواست اضافه شوند
# (جزئیات ناقص → در QA_Notes فهرست می‌شوند تا کاربر کامل کند)
EXTRA_INSTRUMENTS = [
    {"Instrument Code": "FTIR", "Instrument Type": "طیف‌سنج FT-IR",
     "Name FA": "طیف‌سنج مادون قرمز تبدیل فوریه", "Name EN": "FT-IR Spectrometer",
     "Manufacturer / Country": "", "Model": "", "Serial Number": "",
     "Calibration Date": "", "Calibration Interval Months": "12",
     "Location": "آزمایشگاه کنترل کیفیت", "Responsible Person": "", "Status": "Active", "Notes": ""},
]
# نوع/کشور دستگاه‌های اضافه (چون در FIX نیستند)
FIX["FTIR"] = ("", "", "FT-IR Spectrometer", "طیف‌سنج FT-IR")

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb["Instruments"]
rows = list(ws.iter_rows(values_only=True))
headers = [str(h).strip() if h else "" for h in rows[3]]
data = [dict(zip(headers, r)) for r in rows[4:] if any(v not in (None, "") for v in r)]
data += EXTRA_INSTRUMENTS


def g(r, k):
    v = r.get(k)
    return "" if v is None else (v if isinstance(v, datetime) else str(v).strip())

OUT_COLS = ["Instrument Code", "Instrument Type", "Name FA", "Name EN",
            "Manufacturer", "Country", "Model", "Serial Number",
            "Calibration Date", "Calibration Interval Months", "Next Calibration Due",
            "Location", "Responsible Person", "Status", "Notes"]

out_rows = []
qa = []
for r in data:
    code = g(r, "Instrument Code")
    fix = FIX.get(code, ("", "", g(r, "Name EN"), ""))
    manu, country, name_en, itype = fix
    cal = r.get("Calibration Date")
    interval = g(r, "Calibration Interval Months") or "12"
    nxt = add_months(cal, interval) if isinstance(cal, datetime) else ""
    model = MODEL_OVERRIDE.get(code, g(r, "Model"))

    out_rows.append({
        "Instrument Code": code,
        "Instrument Type": itype,
        "Name FA": g(r, "Name FA"),
        "Name EN": name_en,
        "Manufacturer": manu,
        "Country": country,
        "Model": model,
        "Serial Number": g(r, "Serial Number"),
        "Calibration Date": cal if isinstance(cal, datetime) else g(r, "Calibration Date"),
        "Calibration Interval Months": interval,
        "Next Calibration Due": nxt,
        "Location": g(r, "Location"),
        "Responsible Person": g(r, "Responsible Person"),
        "Status": g(r, "Status") or "Active",
        "Notes": g(r, "Notes"),
    })

    if not model:
        qa.append((code, "Model خالی", "مدل دستگاه را وارد کنید"))
    if not manu:
        qa.append((code, "سازنده نامشخص", "نام سازنده را تأیید/کامل کنید"))
    if not country:
        qa.append((code, "کشور نامشخص", "کشور سازنده را وارد کنید"))

# تاریخ کالیبراسیون یکسان تأیید شد (همه در یک روز انجام شده) — دیگر هشدار نیست.

# ---------- نوشتن خروجی ----------
out = openpyxl.Workbook()
sh = out.active
sh.title = "Instruments"

hdr_fill = PatternFill("solid", fgColor="1A3C6E")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, name in enumerate(OUT_COLS, 1):
    cell = sh.cell(1, c, name)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

for ri, row in enumerate(out_rows, 2):
    for ci, name in enumerate(OUT_COLS, 1):
        v = row[name]
        cell = sh.cell(ri, ci, v)
        cell.border = border
        if isinstance(v, datetime):
            cell.number_format = "yyyy-mm-dd"
        if name in ("Instrument Code", "Instrument Type", "Status"):
            cell.font = Font(bold=(name == "Instrument Code"))

widths = [18, 26, 22, 26, 16, 10, 16, 16, 15, 12, 16, 22, 18, 10, 24]
for i, w in enumerate(widths, 1):
    sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
sh.freeze_panes = "A2"

# شیت QA
q = out.create_sheet("QA_Notes")
for c, name in enumerate(["Instrument Code", "مشکل", "اقدام لازم"], 1):
    cell = q.cell(1, c, name); cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center")
for ri, (code, issue, action) in enumerate(qa, 2):
    q.cell(ri, 1, code); q.cell(ri, 2, issue); q.cell(ri, 3, action)
for i, w in enumerate([18, 40, 60], 1):
    q.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
q.freeze_panes = "A2"

out.save(OUT)
print(f"✅ ساخته شد: {OUT}")
print(f"   {len(out_rows)} دستگاه نهایی | {len(qa)} یادداشت QA")
print(f"   ستون‌ها: {len(OUT_COLS)} (شامل Instrument Type و Next Calibration Due جدید)")
