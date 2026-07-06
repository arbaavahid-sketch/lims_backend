"""Build SENAITE Methods and AnalysisService Methods import workbook.

Input:
    TPPC_LIMS_Master_Data_units_filled.xlsx / Method References
    TPPC_AnalysisServices_final.xlsx / Analysis Services

Output:
    TPPC_Methods_Import.xlsx

This keeps the registered-scope traceability from the master workbook while
only linking methods to analysis services that exist in the finalized service
list.
"""
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill


MASTER = "TPPC_LIMS_Master_Data_units_filled.xlsx"
SERVICES = "TPPC_AnalysisServices_final.xlsx"
EXAMPLE = "../../senaite.core/src/bika/lims/setupdata/test/test.xlsx"
OUT = "TPPC_Methods_Import.xlsx"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def kv_from_example(example, sheet, overrides):
    ws = example[sheet]
    keys = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1)) if cell.value]
    records = []
    for row in list(ws.iter_rows(values_only=True))[3:]:
        rec = dict(zip(keys, row))
        field = rec.get("Field")
        if not field:
            continue
        if field in overrides:
            rec["Value"] = overrides[field]
        records.append({key: ("" if rec.get(key) is None else rec.get(key)) for key in keys})
    return records


def method_title(reference_name, reference_year):
    if reference_year:
        return "{} ({})".format(reference_name, reference_year)
    return "{} (year TBD)".format(reference_name)


def main():
    master = openpyxl.load_workbook(MASTER, read_only=True, data_only=True)
    refs_ws = master["Method References"]
    rows = list(refs_ws.iter_rows(values_only=True))
    headers = [clean(value) for value in rows[3]]
    references = [
        dict(zip(headers, row))
        for row in rows[4:]
        if any(value not in (None, "") for value in row)
    ]

    services_wb = openpyxl.load_workbook(SERVICES, read_only=True, data_only=True)
    services_ws = services_wb["Analysis Services"]
    service_rows = list(services_ws.iter_rows(values_only=True))
    header_index = next(
        i for i, row in enumerate(service_rows)
        if row and clean(row[0]) == "Analysis Keyword"
    )
    service_headers = [clean(value) for value in service_rows[header_index]]
    services = [
        dict(zip(service_headers, row))
        for row in service_rows[header_index + 1:]
        if any(value not in (None, "") for value in row)
    ]
    service_title_by_keyword = {
        clean(row.get("Analysis Keyword")): clean(row.get("Analysis Title FA"))
        for row in services
    }

    methods = {}
    method_sources = defaultdict(list)
    links = set()
    qa = []
    for row in references:
        ref_name = clean(row.get("Reference Name"))
        ref_year = clean(row.get("Reference Year"))
        keyword = clean(row.get("Analysis Keyword"))
        if not ref_name:
            continue
        title = method_title(ref_name, ref_year)
        methods[title] = {
            "title": title,
            "description": "Reference: {}".format(ref_name),
            "Instructions": "Year: {}; Clause: {}".format(
                ref_year or "TBD",
                clean(row.get("Reference Clause")) or "TBD",
            ),
            "MethodDocument": "",
            "ManualEntryOfResults": "1",
            "Calculation_title": "",
        }
        method_sources[title].append(clean(row.get("Source Row")))
        service_title = service_title_by_keyword.get(keyword)
        if not service_title:
            qa.append({
                "Issue": "Reference keyword not found in finalized services",
                "Analysis Keyword": keyword,
                "Analysis Title FA": clean(row.get("Analysis Title FA")),
                "Reference Name": ref_name,
                "Reference Year": ref_year,
                "Source Row": clean(row.get("Source Row")),
            })
            continue
        links.add((service_title, title))

    example = openpyxl.load_workbook(EXAMPLE, read_only=True, data_only=True)
    data = {
        "Lab Information": kv_from_example(example, "Lab Information", {
            "Name": "آزمایشگاه پتروشیمی تندیس پارس (TPPC)",
            "LabURL": "https://tppc.ir",
            "Confidence": "95",
            "LaboratoryAccredited": "1",
            "AccreditationBody": "NACI",
            "Physical_Country": "Iran",
            "Postal_Country": "Iran",
            "Billing_Country": "Iran",
        }),
        "Setup": kv_from_example(example, "Setup", {
            "Currency": "IRR",
            "DefaultCountry": "IR",
            "VAT": "9",
            "MemberDiscount": "0",
            "ShowPricing": "0",
            "SamplingWorkflowEnabled": "0",
        }),
        "Methods": list(methods.values()),
        "AnalysisService Methods": [
            {"Service_title": service_title, "Method_title": method}
            for service_title, method in sorted(links)
        ],
    }

    out = openpyxl.Workbook()
    out.remove(out.active)
    header_fill = PatternFill("solid", fgColor="1A3C6E")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_name in example.sheetnames:
        keys = [
            cell.value for cell in next(example[sheet_name].iter_rows(min_row=1, max_row=1))
            if cell.value is not None
        ]
        ws = out.create_sheet(sheet_name[:31])
        for col, key in enumerate(keys, 1):
            cell = ws.cell(1, col, key)
            cell.fill = header_fill
            cell.font = header_font
        ws.cell(2, 1, sheet_name)
        for col, key in enumerate(keys, 1):
            ws.cell(3, col, key)
        for row_index, rec in enumerate(data.get(sheet_name, []), 4):
            for col, key in enumerate(keys, 1):
                ws.cell(row_index, col, rec.get(key, ""))

    qa_ws = out.create_sheet("QA_Notes")
    qa_headers = [
        "Issue", "Analysis Keyword", "Analysis Title FA",
        "Reference Name", "Reference Year", "Source Row",
    ]
    qa_ws.append(qa_headers)
    for cell in qa_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for rec in qa:
        qa_ws.append([rec.get(key, "") for key in qa_headers])
    for col, width in {"A": 44, "B": 18, "C": 54, "D": 24, "E": 14, "F": 12}.items():
        qa_ws.column_dimensions[col].width = width
    qa_ws.freeze_panes = "A2"

    out.save(OUT)
    print("OK created: {}".format(OUT))
    print("methods: {}".format(len(methods)))
    print("service-method links: {}".format(len(links)))
    print("QA rows: {}".format(len(qa)))


if __name__ == "__main__":
    main()
