"""ساخت فایل import افزایشی برای مشخصات (فقط Analysis Specifications).

اسکلت کامل ۶۰ شیت SENAITE (خالی) + Lab Information/Setup (اجباری) +
Analysis Specifications (از TPPC_Specs_Gasoil.xlsx). بقیهٔ داده‌های اپ دست‌نخورده می‌مانند.
خروجی: TPPC_Spec_Import.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

EXAMPLE = "../../senaite.core/src/bika/lims/setupdata/test/test.xlsx"
GASOIL = "TPPC_Specs_Gasoil.xlsx"
OUT = "TPPC_Spec_Import.xlsx"

ex = openpyxl.load_workbook(EXAMPLE, read_only=True, data_only=True)


def kv_from_example(sheet, overrides):
    ws = ex[sheet]
    keys = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value]
    out = []
    for row in list(ws.iter_rows(values_only=True))[3:]:
        rec = dict(zip(keys, row))
        if not rec.get("Field"):
            continue
        if rec["Field"] in overrides:
            rec["Value"] = overrides[rec["Field"]]
        out.append({k: ("" if rec.get(k) is None else rec.get(k)) for k in keys})
    return out


DATA = {}
DATA["Lab Information"] = kv_from_example("Lab Information", {
    "Name": "آزمایشگاه پتروشیمی تندیس پارس (TPPC)", "LabURL": "https://tppc.ir",
    "Confidence": "95", "LaboratoryAccredited": "1", "Accreditation": "ISO/IEC 17025",
    "AccreditationBody": "NACI", "Physical_Country": "Iran",
    "Postal_Country": "Iran", "Billing_Country": "Iran", "EmailAddress": "",
    "Physical_City": "", "Postal_City": "", "Billing_City": "",
    "Physical_Address": "", "Postal_Address": "", "Billing_Address": "",
    "Physical_Zip": "", "Postal_Zip": "", "Billing_Zip": "",
})
DATA["Setup"] = kv_from_example("Setup", {
    "Currency": "IRR", "DefaultCountry": "IR", "VAT": "9",
    "MemberDiscount": "0", "ShowPricing": "0", "SamplingWorkflowEnabled": "0",
})

# Analysis Specifications را از فایل گازوئیل بخوان (ردیف ۴ به بعد)
gw = openpyxl.load_workbook(GASOIL, data_only=True)["Analysis Specifications"]
grows = list(gw.iter_rows(values_only=True))
gkeys = [c for c in grows[0] if c]
specs = []
for row in grows[3:]:
    rec = dict(zip(gkeys, row))
    if rec.get("SampleType_title") and rec.get("service"):
        specs.append({k: ("" if rec.get(k) is None else rec.get(k)) for k in gkeys})
DATA["Analysis Specifications"] = specs

# ---------- نوشتن اسکلت کامل ----------
out = openpyxl.Workbook(); out.remove(out.active)
hf = PatternFill("solid", fgColor="1A3C6E"); hfont = Font(color="FFFFFF", bold=True)
for name in ex.sheetnames:
    keys = [c.value for c in next(ex[name].iter_rows(min_row=1, max_row=1)) if c.value is not None]
    ws = out.create_sheet(name[:31])
    for c, k in enumerate(keys, 1):
        cell = ws.cell(1, c, k); cell.fill = hf; cell.font = hfont
    ws.cell(2, 1, name)
    for c, k in enumerate(keys, 1):
        ws.cell(3, c, k)
    for ri, rec in enumerate(DATA.get(name, []), 4):
        for ci, k in enumerate(keys, 1):
            ws.cell(ri, ci, rec.get(k, ""))

out.save(OUT)
print(f"✅ {OUT} | Analysis Specifications: {len(specs)} ردیف")
for s in specs:
    print(f"   {s['SampleType_title']} | {str(s['service'])[:34]:34} | min={s['min']} max={s['max']}")
