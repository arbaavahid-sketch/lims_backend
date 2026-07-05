"""import افزایشی خانواده‌های باقی‌مانده: ضدیخ، گریس، روغن موتور بنزینی/دیزلی.

منبع: Found_Limits_v2 در TPPC_Test_Limits_Master_v2_completed.xlsx
خروجی: TPPC_Families_Import.xlsx (افزایشی — گرید=نوع‌نمونه، پارامتر=سرویس، حد=مشخصات)

نکتهٔ صحت: چون importer حد خالی را «۰» می‌کند و کد conformity ۰ را literal می‌گیرد،
از سنتینل استفاده می‌کنیم: بی‌حدِ پایین = -999999، بی‌حدِ بالا = +999999.
این برای پارامترهای منفی‌پذیر (نقطه انجماد) حیاتی است.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

SRC = r"G:\Downloads\TPPC_Test_Limits_Master_v2_completed.xlsx"
EXAMPLE = "../../senaite.core/src/bika/lims/setupdata/test/test.xlsx"
OUT = "TPPC_Families_Import.xlsx"
NEG = -999999
POS = 999999

# خانواده → (آزمایشگاه/دپارتمان، پیشوند عنوان نوع‌نمونه یا None اگر گرید خودش گویاست)
FAMILIES = {
    "ضدیخ": ("ضدیخ", None),
    "گریس": ("گریس", None),
    "روغن موتور بنزینی": ("روغن", "روغن بنزینی "),
    "روغن موتور دیزلی": ("روغن", "روغن دیزلی "),
}

wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb["Found_Limits_v2"]
rows = list(ws.iter_rows(values_only=True))
hi = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Limit_ID")
h = [str(x).strip() if x else "" for x in rows[hi]]
data = [dict(zip(h, r)) for r in rows[hi+1:] if any(v not in (None, "") for v in r)]


def gv(r, k):
    v = r.get(k); s = "" if v is None else str(v).strip()
    return "" if s == "." else s


def num(x):
    try:
        return float(x)
    except (ValueError, TypeError):
        return None


sample_types = {}      # title -> prefix
services = {}          # param -> (keyword, unit, lab)
specs = []
svc_i = 0
skipped = 0

for r in data:
    fam = gv(r, "Product_Family")
    if fam not in FAMILIES:
        continue
    op = gv(r, "Limit_Operator")
    if op not in ("Min", "Max", "Range"):
        skipped += 1
        continue
    mn, mx = num(gv(r, "Min_Value")), num(gv(r, "Max_Value"))
    if mn is None and mx is None:
        skipped += 1
        continue
    # «مثال قراردادی» و مشابه را رد کن
    param = gv(r, "Parameter")
    if "مثال" in param or "قرارداد" in param:
        skipped += 1
        continue

    lab, prefix = FAMILIES[fam]
    grade = gv(r, "Product/Grade")
    st_title = (prefix + grade) if prefix else grade
    if st_title not in sample_types:
        sample_types[st_title] = f"ST{len(sample_types)+1:02d}F"

    if param not in services:
        svc_i += 1
        services[param] = (f"SPX_{svc_i:03d}", gv(r, "Unit"), lab)

    # سنتینل برای سمت بی‌حد
    smin = mn if mn is not None else NEG
    smax = mx if mx is not None else POS
    specs.append({"Title": st_title, "SampleType_title": st_title,
                  "service": param, "min": smin, "max": smax})

# ---------- ساخت فایل import (اسکلت + LabInfo/Setup + ۳ شیت) ----------
ex = openpyxl.load_workbook(EXAMPLE, read_only=True, data_only=True)


def kv(sheet, ov):
    wsx = ex[sheet]
    keys = [c.value for c in next(wsx.iter_rows(min_row=1, max_row=1)) if c.value]
    out = []
    for row in list(wsx.iter_rows(values_only=True))[3:]:
        rec = dict(zip(keys, row))
        if not rec.get("Field"):
            continue
        if rec["Field"] in ov:
            rec["Value"] = ov[rec["Field"]]
        out.append({k: ("" if rec.get(k) is None else rec.get(k)) for k in keys})
    return out


DATA = {
    "Lab Information": kv("Lab Information", {
        "Name": "آزمایشگاه پتروشیمی تندیس پارس (TPPC)", "LabURL": "https://tppc.ir",
        "Confidence": "95", "LaboratoryAccredited": "1", "Accreditation": "ISO/IEC 17025",
        "AccreditationBody": "NACI", "Physical_Country": "Iran", "Postal_Country": "Iran",
        "Billing_Country": "Iran", "EmailAddress": "", "Physical_City": "", "Postal_City": "",
        "Billing_City": "", "Physical_Address": "", "Postal_Address": "", "Billing_Address": "",
        "Physical_Zip": "", "Postal_Zip": "", "Billing_Zip": ""}),
    "Setup": kv("Setup", {"Currency": "IRR", "DefaultCountry": "IR", "VAT": "9",
                          "MemberDiscount": "0", "ShowPricing": "0", "SamplingWorkflowEnabled": "0"}),
    "Sample Types": [{"title": t, "Prefix": p, "MinimumVolume": "0 mL",
                      "description": "", "Hazardous": "0"} for t, p in sample_types.items()],
    "Analysis Services": [{"title": param, "Keyword": kw, "PointOfCapture": "lab",
                           "AnalysisCategory_title": lab, "Department_title": lab, "Unit": unit,
                           "ManualEntryOfResults": "1", "Accredited": "1", "Price": "0"}
                          for param, (kw, unit, lab) in services.items()],
    "Analysis Specifications": [{"Title": s["Title"], "Client_title": "",
                                 "SampleType_title": s["SampleType_title"], "service": s["service"],
                                 "min": s["min"], "max": s["max"], "error": ""} for s in specs],
}

out = openpyxl.Workbook(); out.remove(out.active)
hf = PatternFill("solid", fgColor="1A3C6E"); hfont = Font(color="FFFFFF", bold=True)
for name in ex.sheetnames:
    keys = [c.value for c in next(ex[name].iter_rows(min_row=1, max_row=1)) if c.value is not None]
    wso = out.create_sheet(name[:31])
    for c, k in enumerate(keys, 1):
        cell = wso.cell(1, c, k); cell.fill = hf; cell.font = hfont
    wso.cell(2, 1, name)
    for c, k in enumerate(keys, 1):
        wso.cell(3, c, k)
    for ri, rec in enumerate(DATA.get(name, []), 4):
        for ci, k in enumerate(keys, 1):
            wso.cell(ri, ci, rec.get(k, ""))
out.save(OUT)
print(f"✅ {OUT}")
print(f"   نوع‌نمونه (گرید): {len(sample_types)}")
print(f"   سرویس آزمون: {len(services)}")
print(f"   ردیف مشخصات: {len(specs)} | رد شده (متنی/خالی): {skipped}")
print("\n   سرویس‌های ساخته‌شده:")
for param, (kw, unit, lab) in services.items():
    print(f"     [{lab:5}] {param[:36]:36} ({unit})")
