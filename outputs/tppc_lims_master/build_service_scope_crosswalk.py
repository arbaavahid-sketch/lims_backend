from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SERVICES_PATH = ROOT / "outputs" / "tppc_lims_master" / "TPPC_AnalysisServices_final.xlsx"
OUT_PATH = ROOT / "docs" / "PHASE1_SERVICE_SCOPE_CROSSWALK.md"
SOURCE_PDF = r"D:\Documents\SortedDocs\PDF\شرح خدمات آزمایشگاه روغن.pdf"


@dataclass(frozen=True)
class ScopeRow:
    group: str
    no: str
    title: str
    refs: tuple[str, ...]
    decision: str
    note: str


ROWS: tuple[ScopeRow, ...] = (
    ScopeRow("Grease", "1", "Grease corrosion", ("ASTM D 4048", "INSO 11291"), "Covered", "Covered by the current copper/grease corrosion service."),
    ScopeRow("Grease", "2", "Flash point on separated grease oil", ("ASTM D 92", "INSO 198"), "Review", "INSO 198 exists, but the current method set is mostly Pensky-Martens/ASTM D93. Confirm Cleveland open-cup D92 service."),
    ScopeRow("Grease", "3", "Atmospheric distillation on separated grease oil", ("ASTM D 86", "INSO 6261"), "Covered", "Covered by atmospheric distillation services."),
    ScopeRow("Grease", "4", "Water content", ("ASTM D4048", "INSO 11291"), "Review", "Source row likely has a method/title mismatch; ASTM D4048/INSO11291 are corrosion references, not water content."),
    ScopeRow("Bitumen", "1", "Bitumen ductility", ("ASTM D 113", "INSO 3866"), "Add/Review", "No direct service found. Add only if bitumen ductility is in active lab scope."),
    ScopeRow("Bitumen", "2", "Bitumen specific gravity", ("ASTM D 70", "INSO 3872"), "Add/Review", "No direct bitumen D70/INSO3872 service found."),
    ScopeRow("Bitumen", "3", "Bitumen viscosity", ("ASTM D 445", "INSO 340"), "Review", "Viscosity services exist; confirm product scope includes bitumen for this method."),
    ScopeRow("Bitumen", "4", "Extraction solvent distillation", ("ASTM D 86",), "Review", "D86 exists; confirm this bitumen/extraction-solvent scope is intended."),
    ScopeRow("Bitumen", "5", "Spot test", ("AASHTO T102-83", "ISIRI 2949"), "Covered", "Covered by existing bitumen spot-test service."),
    ScopeRow("Antifreeze", "1", "Antifreeze pH", ("ASTM D 1287", "INSO 1212"), "Add/Review", "No direct antifreeze pH service found."),
    ScopeRow("Antifreeze", "2", "Freezing point", ("ASTM D 1177", "INSO 1448"), "Add/Review", "No direct antifreeze freezing-point service found."),
    ScopeRow("Antifreeze", "3", "Relative density", ("ASTM D 1122",), "Deferred/Review", "PDF marks this as future-capable; do not import unless active scope is confirmed."),
    ScopeRow("Antifreeze", "4", "Water content", ("ASTM D 1123", "ISIRI 6228"), "Add/Review", "No direct ASTM D1123/ISIRI6228 service found."),
    ScopeRow("Antifreeze", "6", "Chloride", ("ASTM D 3634", "ISIRI 5596"), "Add/Review", "Current chloride service uses ASTM E2469 for ethylene glycol; confirm whether D3634/ISIRI5596 is required."),
    ScopeRow("Oil", "1", "Color", ("ASTM D 1500", "INSO 203"), "Covered", "Covered by color services."),
    ScopeRow("Oil", "2", "Flash point open cup", ("ASTM D 92", "INSO 198"), "Review", "INSO 198 exists, but ASTM D92/open-cup should be verified separately from D93 closed-cup."),
    ScopeRow("Oil", "3", "Pour point", ("ASTM D 97", "INSO 201"), "Covered", "Covered by pour-point services."),
    ScopeRow("Oil", "4", "Viscosity", ("ASTM D 445", "INSO 195"), "Covered", "Covered by viscosity and viscosity-index services."),
    ScopeRow("Oil", "5", "Copper strip corrosion", ("ASTM D 130", "ISIRI 336"), "Covered", "Covered by copper-corrosion/sulfur method references, but ISIRI336 usage should be checked in final method mapping."),
    ScopeRow("Oil", "6", "Total sulfur", ("ASTM D 4294", "INSO 8402"), "Covered", "Covered by XRF sulfur service."),
    ScopeRow("Oil", "7", "Ash", ("ASTM D 482", "INSO 2940"), "Covered", "Covered by ash services."),
    ScopeRow("Oil", "8", "TBN", ("ASTM D 2896", "INSO 2772"), "Covered", "Covered by TBN service."),
    ScopeRow("Oil", "9", "TAN", ("ASTM D 664", "INSO 18030"), "Covered", "Covered by TAN service."),
    ScopeRow("Oil", "10", "Water by Karl Fischer", ("ASTM D 6304", "INSO 18481"), "Covered", "ASTM D6304 services exist; confirm whether INSO18481 must be attached to the same service."),
    ScopeRow("Oil", "11", "Elemental analysis, oil, 19 elements", ("ASTM D 6595",), "Review", "ICP element services exist, but ASTM D6595 as a single 19-element oil package is not explicit."),
    ScopeRow("Oil", "12", "Water separability", ("ASTM D1401", "INSO 3169"), "Review", "Current water-separability service uses ISO 6614. Confirm D1401/INSO3169 method mapping."),
    ScopeRow("Oil", "13", "Ramsbottom carbon residue", ("ASTM D524", "INSO 200"), "Add/Review", "No direct carbon-residue service found."),
    ScopeRow("Oil", "14", "Corrosion at 135 C", ("ASTM D6594",), "Covered", "Covered by the current 135 C corrosion service."),
    ScopeRow("Oil", "15", "Corrosion at 125 C", ("ASTM D5968",), "Covered", "Covered after notation cleanup; current source uses ASTM 5968 without D."),
    ScopeRow("Hydrocarbon", "1", "Color", ("ASTM D1500-156", "INSO 203", "INSO 2932"), "Covered", "Color services exist; confirm dual D1500/D156 notation if this is a combined row."),
    ScopeRow("Hydrocarbon", "2", "Flash point closed cup", ("ASTM D 93", "INSO 19695"), "Covered", "Covered by closed-cup flash-point services."),
    ScopeRow("Hydrocarbon", "3", "Flash point open cup", ("ASTM D 92", "INSO 198"), "Review", "Open-cup D92 should be verified separately from D93 closed-cup."),
    ScopeRow("Hydrocarbon", "4", "Atmospheric distillation", ("ASTM D 86", "INSO 6261"), "Covered", "Covered by D86 distillation services."),
    ScopeRow("Hydrocarbon", "5", "Cloud point", ("ASTM D 2500", "INSO 5438"), "Covered", "Covered by cloud-point service."),
    ScopeRow("Hydrocarbon", "6", "Pour point", ("ASTM D 97", "INSO 201"), "Covered", "Covered by pour-point services."),
    ScopeRow("Hydrocarbon", "7", "Viscosity", ("ASTM D 445", "ISIRI 340"), "Covered", "Covered by viscosity services; normalize ISIRI/INSO 340 naming."),
    ScopeRow("Hydrocarbon", "8", "Copper strip corrosion", ("ASTM D 130", "ISIRI 336"), "Covered", "Covered by copper-corrosion services."),
    ScopeRow("Hydrocarbon", "9", "Total sulfur", ("ASTM D 4294", "INSO 8402"), "Covered", "Covered by XRF sulfur service."),
    ScopeRow("Hydrocarbon", "10", "Mercaptan", ("ASTM D 3227", "INSO 9379"), "Covered", "Covered by mercaptan sulfur services."),
    ScopeRow("Hydrocarbon", "11", "Hydrogen sulfide", ("ASTM D 3227", "INSO 9379"), "Review", "Method reference overlaps mercaptan; title is hydrogen sulfide. Confirm actual method/service."),
    ScopeRow("Hydrocarbon", "12", "Chlorine", ("ISO 15597", "ISIRI 13378"), "Add/Review", "No direct ISO15597/ISIRI13378 chlorine service found."),
    ScopeRow("Hydrocarbon", "13", "Doctor test", ("ASTM D 4952", "ISIRI 8722"), "Add/Review", "No direct doctor-test service found."),
    ScopeRow("Hydrocarbon", "14", "Ash", ("ASTM D 482", "INSO 2940"), "Covered", "Covered by ash services."),
    ScopeRow("Hydrocarbon", "15", "Water by distillation", ("ASTM D 95", "ISIRI 8139", "ISIRI 4081"), "Covered", "Covered by water-distillation service."),
    ScopeRow("Hydrocarbon", "16", "Water by Karl Fischer", ("ASTM D 1796", "INSO 18481"), "Review", "Title says Karl Fischer, but ASTM D1796 is centrifuge-style water/sediment. Confirm/correct source row."),
    ScopeRow("Hydrocarbon", "17", "Elemental analysis, oil, 19 elements", ("ASTM D 6595",), "Review", "ICP element services exist, but ASTM D6595 package is not explicit."),
    ScopeRow("Hydrocarbon", "18", "GC benzene/aromatic/olefin", ("ASTM D 6730", "ASTM D 5134", "ASTM D 6729"), "Review", "D6730 exists; D5134/D6729 are not explicit in current service references."),
    ScopeRow("Hydrocarbon", "19", "PONA", ("ASTM D 1319-GC", "ISIRI 8403"), "Review", "D1319 reference appears, but PONA/GC scope and ISIRI8403 should be confirmed."),
    ScopeRow("Hydrocarbon", "20", "Sediment by extraction", ("ASTM D473", "INSO 1210", "ISIRI 4188"), "Covered", "ASTM D473 sediment service exists; INSO1210/ISIRI4188 need method-link review."),
    ScopeRow("Crude Oil", "1", "Specific gravity at 15.6 C", ("ASTM A1298", "INSO 197"), "Review", "Likely source typo for ASTM D1298; D1298/INSO197 service exists."),
    ScopeRow("Crude Oil", "2", "API Gravity", ("ASTM D 1298", "INZO 197"), "Covered", "Covered by D1298/INSO197 after correcting INZO typo to INSO."),
    ScopeRow("Crude Oil", "3", "Total sulfur", ("ASTM D 4294", "INSO 8402"), "Covered", "Covered by XRF sulfur service."),
    ScopeRow("Crude Oil", "4", "Mercaptan", ("ASTM D 3227", "INSO 9379"), "Covered", "Covered by mercaptan sulfur services."),
    ScopeRow("Crude Oil", "5", "Water and sediment by centrifuge", ("ASTM D4007", "INSO 15342"), "Covered", "ASTM D4007 service exists; INSO15342 should be added/checked in method links."),
    ScopeRow("Crude Oil", "6", "Water by distillation", ("ISO 9029", "ISIRI 8139"), "Review", "Water-distillation service exists via ASTM D95/ISIRI8139; ISO9029 is not explicit."),
    ScopeRow("Crude Oil", "7", "Viscosity", ("ASTM D 445", "ISIRI 340"), "Covered", "Covered by viscosity services; normalize ISIRI/INSO 340 naming."),
    ScopeRow("Crude Oil", "8", "Pour point", ("ASTM D 97", "INSO 201"), "Covered", "Covered by pour-point services."),
    ScopeRow("Crude Oil", "9", "Vapor pressure", ("ASTM D 323", "INSO 5439"), "Covered", "Covered by RVP/vapor-pressure service."),
    ScopeRow("Crude Oil", "10", "Ramsbottom carbon residue", ("ASTM D5424", "INSO 200"), "Add/Review", "No direct carbon-residue service found. Also confirm whether PDF meant ASTM D524."),
    ScopeRow("Crude Oil", "11", "TAN", ("ASTM D 664", "INSO 18030"), "Covered", "Covered by TAN service."),
    ScopeRow("Crude Oil", "12", "Atmospheric distillation", ("ASTM D 86", "INSO 6261"), "Covered", "Covered by D86 distillation services."),
)


def norm(value: str) -> str:
    value = value.upper().replace("INZO", "INSO")
    return re.sub(r"[^A-Z0-9]+", "", value)


def relaxed_ref_keys(ref: str) -> set[str]:
    keys = {norm(ref)}
    m = re.search(r"ASTM\s*D\s*(\d+)\s*-\s*(\d+)", ref, flags=re.I)
    if m:
        keys.add(f"ASTMD{m.group(1)}")
        keys.add(f"ASTMD{m.group(2)}")
    m = re.search(r"ASTM\s*D\s*(\d+)", ref, flags=re.I)
    if m:
        keys.add(f"ASTM{m.group(1)}")
        keys.add(f"ASTMD{m.group(1)}")
    m = re.search(r"ASTM\s*A\s*(\d+)", ref, flags=re.I)
    if m:
        keys.add(f"ASTMD{m.group(1)}")
    return keys


def extract_ref_keys(refs: str) -> set[str]:
    keys: set[str] = set()
    patterns = (
        r"ASTM\s*D\s*\d+",
        r"ASTM\s*A\s*\d+",
        r"ASTM\s+\d+",
        r"AASHTO\s*T\s*\d+(?:-\d+)?",
        r"A\.A\.S\.H\.T\.O\s*T\s*\d+(?:-\d+)?",
        r"ISO\s*\d+",
        r"INSO\s*\d+",
        r"ISIRI\s*\d+",
        r"INZO\s*\d+",
        r"IEC\s*\d+",
    )
    for pattern in patterns:
        for match in re.finditer(pattern, refs, flags=re.I):
            keys.update(relaxed_ref_keys(match.group(0)))
    return keys


def load_services() -> list[dict[str, str]]:
    workbook = load_workbook(SERVICES_PATH, read_only=True, data_only=True)
    sheet = workbook["Analysis Services"]
    services: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        keyword, title, _unit, _dept, instrument, _tat, refs, _count, products, *_rest = row
        services.append(
            {
                "keyword": str(keyword or ""),
                "title": str(title or ""),
                "refs": str(refs or ""),
                "instrument": str(instrument or ""),
                "products": str(products or ""),
                "refs_norm": norm(str(refs or "")),
                "ref_keys": extract_ref_keys(str(refs or "")),
            }
        )
    return services


def matching_services(scope_row: ScopeRow, services: list[dict[str, str]]) -> list[dict[str, str]]:
    keys = set()
    for ref in scope_row.refs:
        keys.update(relaxed_ref_keys(ref))
    matches = []
    for service in services:
        if keys.intersection(service["ref_keys"]):
            matches.append(service)
    return matches


def summarize_matches(matches: list[dict[str, str]]) -> str:
    if not matches:
        return "-"
    chunks = []
    for service in matches[:4]:
        chunks.append(f"{service['keyword']} - {service['title']}")
    if len(matches) > 4:
        chunks.append(f"... +{len(matches) - 4} more")
    return "<br>".join(chunks)


def main() -> None:
    services = load_services()
    rows = []
    for row in ROWS:
        matches = matching_services(row, services)
        rows.append((row, matches))

    counts: dict[str, int] = {}
    for row, _matches in rows:
        counts[row.decision] = counts.get(row.decision, 0) + 1

    lines = [
        "# Phase 1 Service Scope Crosswalk - Oil Lab PDF",
        "",
        f"Source PDF: `{SOURCE_PDF}`",
        "",
        "Generated on: 2026-07-06",
        "",
        "## Summary",
        "",
        f"- PDF service rows reviewed: {len(rows)}",
        f"- Covered: {counts.get('Covered', 0)}",
        f"- Review: {counts.get('Review', 0)}",
        f"- Add/Review: {counts.get('Add/Review', 0)}",
        f"- Deferred/Review: {counts.get('Deferred/Review', 0)}",
        "",
        "## Decision Rules",
        "",
        "- `Covered`: a current Analysis Service appears to cover the row after normalizing method notation.",
        "- `Review`: current data is close, but method, product scope, package grouping, or source typo needs a human decision.",
        "- `Add/Review`: no sufficient current Analysis Service was found; add only if the lab confirms the service is active.",
        "- `Deferred/Review`: source says future-capable or otherwise not ready for import.",
        "",
        "## Crosswalk",
        "",
        "| Group | PDF Row | PDF Test | PDF Method References | Decision | Candidate Current Services | Note |",
        "| --- | ---: | --- | --- | --- | --- | --- |",
    ]

    for row, matches in rows:
        refs = "; ".join(row.refs)
        lines.append(
            f"| {row.group} | {row.no} | {row.title} | {refs} | {row.decision} | "
            f"{summarize_matches(matches)} | {row.note} |"
        )

    lines.extend(
        [
            "",
            "## Next Actions",
            "",
            "1. Resolve all `Add/Review` rows with the lab technical owner.",
            "2. Resolve `Review` rows by either normalizing method references, expanding product scope, or marking the PDF row as source-only evidence.",
            "3. Do not import new Analysis Services until method document links and responsible instruments/people are known.",
        ]
    )

    OUT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_PATH}")
    print(f"Rows: {len(rows)}")
    for key in ("Covered", "Review", "Add/Review", "Deferred/Review"):
        print(f"{key}: {counts.get(key, 0)}")


if __name__ == "__main__":
    main()
