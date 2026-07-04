"""نگاشت خودکار Analysis Services به دستگاه/دپارتمان/زمان انجام.

بر اساس نام فارسی هر آزمون، دستگاهِ مناسب از فهرست ۲۹تایی دستگاه‌ها انتخاب می‌شود.
مواردی که با اطمینان کافی نگاشت نشوند، در ستون Notes با «REVIEW» علامت می‌خورند.

ورودی: TPPC_LIMS_Master_Data_units_filled.xlsx (شیت Analysis Services)
خروجی: TPPC_AnalysisServices_final.xlsx
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "TPPC_LIMS_Master_Data_units_filled.xlsx"
OUT = "TPPC_AnalysisServices_final.xlsx"

# قواعد نگاشت: (الگوی نام فارسی, کد دستگاه, دپارتمان, زمان‌روز)
# ترتیب مهم است — اولین تطبیق برنده است، پس خاص‌ها را بالاتر بگذار.
RULES = [
    (r"کارل\s*فیشر.*کولوم|کولوم.*کارل|کولومتری", "Coulometry", "Instrumental", 1),
    (r"کارل\s*فیشر|دین\s*استارک", "KARL_FISHER", "Oil Lab", 1),
    (r"گوگرد|سولفور|سولفات", "SE", "Instrumental", 1),
    (r"گرانرو|ویسکوز|ویسکو|رانش|کینماتیک", "VM", "Oil Lab", 1),
    (r"نقطه\s*اشتعال|اشتعال|فلش", "FP_01", "Oil Lab", 1),
    (r"چگالی|دانسیت|API|وزن\s*مخصوص|دانسیته", "Densitimeter", "Oil Lab", 1),
    (r"تقطیر", "Distilation_01", "Oil Lab", 1),
    (r"نقطه\s*ریزش|ریزش", "Pour_point", "Oil Lab", 1),
    (r"نقطه\s*ابری|ابری", "Cloud_point", "Oil Lab", 1),
    (r"اکتان", "OCTAN_CETAN_METER", "Oil Lab", 1),
    (r"ستان", "OCTAN_CETAN_METER", "Oil Lab", 1),
    (r"فشار\s*بخار|رید|RVP", "RVP", "Oil Lab", 1),
    (r"خوردگی.*مس|نوار\s*مس|خوردگی", "CS", "Oil Lab", 1),
    (r"عدد\s*اسید|اسید.*عدد|اسیدیت.*روغن|خنثی\s*ساز|عدد\s*قلیای|قلیای|بازی", "PT", "Oil Lab", 1),
    (r"هدایت|کنداکت|رسانای", "COND", "Water Lab", 1),
    (r"کدورت", "Turbidity", "Water Lab", 1),
    (r"pH|اسیدیته|پ\s*هاش|پ‌هاش", "PH", "Water Lab", 1),
    (r"رنگ|سیبولت|لاوی|saybolt", "saybolt", "Oil Lab", 1),
    (r"کروماتوگراف|بنزین.*آنالیز|DHA|هیدروکرب|ترکیب.*بنزین", "GC_DHA", "Instrumental", 2),
    (r"آلومینیوم|باریم|بور|تیتانیوم|استرانسیم|کلسیم|سدیم|پتاسیم|وانادیوم|نیکل|"
     r"آهن\b|روی\b|فسفر|منیزیم|سرب|سیلیک|منگنز|مولیبد|کبالت|قلع|نقره|عناصر|فلز|"
     r"طیف.*نشر.*پلاسما|ICP", "ICP", "Instrumental", 2),
    (r"عبور\s*UV|UV|فرابنفش|جذب\s*نور|اسپکتروف|رنگ\s*سنجی|فتومتر", "DR", "Instrumental", 1),
    (r"نامحلول|رسوب|خاکستر|وزن.*باقی|گراویمتر", "INST_027", "Oil Lab", 1),
]

# دپارتمان پیش‌فرض برای موارد نگاشت‌نشده
DEFAULT_DEPT = "Oil Lab"

# سنتینل: آزمون‌هایی که ذاتاً دستگاه ندارند (دستی/بصری/محاسباتی/نمونه‌برداری)
NONE = "-"

# نگاشت صریح موارد خاص (زیررشته در نام فارسی → کد دستگاه/دپارتمان/زمان/یادداشت)
# NONE یعنی «بدون دستگاه» و خطا محسوب نمی‌شود؛ "" یعنی هنوز نیاز به تصمیم کاربر (REVIEW).
OVERRIDES = [
    ("قابلیت جدا شدن آب", NONE, "Oil Lab", 1, "حمام جداسازی آب/روغن (دستی)"),
    ("لکه مواد قیری", NONE, "Oil Lab", 1, "آزمون بصری"),
    ("آماده سازی نمونه", NONE, "Oil Lab", 1, "روال آماده‌سازی — آزمون نیست"),
    ("نمونه برداری دستی", NONE, "Oil Lab", 1, "نمونه‌برداری — آزمون نیست"),
    ("اسید آزاد", "PT", "Oil Lab", 1, ""),
    ("اندازه گیری روغن", "INST_027", "Oil Lab", 1, "توزین/استخراج"),
    ("قلیائى کل", "PT", "Oil Lab", 1, "تیتراسیون (TBN)"),
    ("مرکاپتان", "PT", "Oil Lab", 1, "تیتراسیون پتانسیومتری"),
    ("مقدار آب", "KARL_FISHER", "Oil Lab", 1, ""),
    ("آلدهید", "TITRATOR", "Oil Lab", 1, ""),
    ("کلراید", "PT", "Water Lab", 1, "تیتراسیون"),
    ("Specific Gravity در گلایکول", "Densitimeter", "Oil Lab", 1, ""),
    ("گرمای احتراق", NONE, "Oil Lab", 1, "کالریمتر بمبی موجود نیست / محاسباتی"),
    ("حلالیت مواد قیری", "INST_027", "Oil Lab", 1, "توزین"),
    ("ذرات سخت", "INST_027", "Oil Lab", 1, "توزین/صافش"),
    ("نقطه جوش", "Distilation_01", "Oil Lab", 1, ""),
    ("خنثی شدن کل", "PT", "Oil Lab", 1, "تیتراسیون"),
    ("n-d-M", NONE, "Oil Lab", 1, "محاسباتی از n، d، M"),
    ("کراکل", "HEATER_STIRRER", "Oil Lab", 1, ""),
    ("TAME", "GC_DHA", "Instrumental", 2, "کروماتوگرافی گازی"),
    ("FT-IR", "FTIR", "Instrumental", 2, ""),
    ("مادون قرمز", "FTIR", "Instrumental", 2, ""),
    ("افزودنی های بازدارنده", "FTIR", "Oil Lab", 1, "کمی‌سازی با FT-IR — تأیید کنید"),
]


def normalize(s):
    # یکسان‌سازی «ی/ك» عربی و فاصله‌ها برای تطبیق مطمئن‌تر
    return (s.replace("ي", "ی").replace("ى", "ی").replace("ك", "ک"))


def match(title):
    t = normalize(title)
    for key, inst, dept, tat, note in OVERRIDES:
        if normalize(key) in t:
            return inst, dept, tat, note
    for pat, inst, dept, tat in RULES:
        if re.search(normalize(pat), t):
            return inst, dept, tat, ""
    return "", DEFAULT_DEPT, "", ""


wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb["Analysis Services"]
rows = list(ws.iter_rows(values_only=True))
hdr_i = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Analysis Keyword")
headers = [str(h).strip() if h else "" for h in rows[hdr_i]]
data = [dict(zip(headers, r)) for r in rows[hdr_i+1:] if any(v not in (None, "") for v in r)]


def g(r, k):
    v = r.get(k)
    return "" if v is None else str(v).strip()

matched = review = no_inst = 0
for r in data:
    title = g(r, "Analysis Title FA")
    inst, dept, tat, note = match(title)
    r["Department"] = dept
    r["Turnaround Days"] = tat
    if inst == NONE:            # آزمون بدون دستگاه (دستی/محاسباتی) — درست است، خطا نیست
        r["Instrument Code"] = ""
        r["Notes"] = "بدون دستگاه — " + note
        no_inst += 1
    elif inst == "":           # واقعاً نامشخص — نیاز به تصمیم کاربر
        r["Instrument Code"] = ""
        r["Notes"] = "REVIEW: " + (note or "دستگاه نامشخص — دستی تعیین کنید")
        review += 1
    else:                      # نگاشت موفق
        r["Instrument Code"] = inst
        r["Notes"] = note
        matched += 1

# ---------- نوشتن خروجی ----------
out = openpyxl.Workbook()
sh = out.active
sh.title = "Analysis Services"
COLS = headers  # همان ستون‌ها، با ۳ ستون پرشده

hdr_fill = PatternFill("solid", fgColor="1A3C6E")
hdr_font = Font(color="FFFFFF", bold=True, size=10)
rev_fill = PatternFill("solid", fgColor="FDECEA")  # قرمز روشن برای REVIEW
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, name in enumerate(COLS, 1):
    cell = sh.cell(1, c, name)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for ri, r in enumerate(data, 2):
    is_rev = str(r.get("Notes") or "").startswith("REVIEW")
    for ci, name in enumerate(COLS, 1):
        cell = sh.cell(ri, ci, r.get(name))
        cell.border = border
        if is_rev and name in ("Instrument Code", "Notes"):
            cell.fill = rev_fill

widths = {"Analysis Keyword": 16, "Analysis Title FA": 42, "Unit": 14, "Department": 14,
          "Instrument Code": 18, "Turnaround Days": 10, "Method References": 20,
          "Example Products": 24, "Notes": 40}
for i, name in enumerate(COLS, 1):
    sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths.get(name, 12)
sh.freeze_panes = "A2"

out.save(OUT)
print(f"✅ ساخته شد: {OUT}")
print(f"   کل: {len(data)} | نگاشت‌شده: {matched} | بدون دستگاه (درست): {no_inst} | نیاز به بازبینی: {review}")

# توزیع دستگاه‌ها
from collections import Counter
dist = Counter(g(r, "Instrument Code") or "(REVIEW)" for r in data)
print("\n   توزیع دستگاه:")
for inst, n in dist.most_common():
    print(f"     {inst:20} {n}")
