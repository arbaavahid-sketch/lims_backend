"""import افزایشی کاربران (Lab Contacts + نقش) برای تفکیک وظایف ISO 17025 بند 6.2.

خروجی: TPPC_Users_Import.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

EXAMPLE = "../../senaite.core/src/bika/lims/setupdata/test/test.xlsx"
OUT = "TPPC_Users_Import.xlsx"
PW = "Tandis1403"

# (نام، نام‌خانوادگی، username, نقش‌ها, گروه‌ها, دپارتمان, سمت)
USERS = [
    ("منصور", "آزمایش‌گر", "analyst",  "Analyst",    "Analysts",    "هیدروکربن", "کارشناس آزمون"),
    ("سمانه", "نمونه‌بردار", "sampler",  "Sampler",    "Samplers",    "نفت خام",   "نمونه‌بردار"),
    ("دکتر",  "مدیرفنی",   "manager",  "LabManager,Publisher", "LabManagers", "هیدروکربن", "مدیر فنی"),
    # بازبین باید هم نمونه را ببیند (Analyst) هم تأیید کند (Verifier) —
    # نقش Verifier به‌تنهایی اجازهٔ مشاهدهٔ نمونه را نمی‌دهد.
    ("رضا",   "بازبین",    "reviewer", "Verifier,Analyst", "Verifiers,Analysts", "روغن", "بازبین نتایج"),
]

ex = openpyxl.load_workbook(EXAMPLE, read_only=True, data_only=True)


def kv(sheet, ov):
    ws = ex[sheet]
    keys = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value]
    out = []
    for row in list(ws.iter_rows(values_only=True))[3:]:
        rec = dict(zip(keys, row))
        if not rec.get("Field"):
            continue
        if rec["Field"] in ov:
            rec["Value"] = ov[rec["Field"]]
        out.append({k: ("" if rec.get(k) is None else rec.get(k)) for k in keys})
    return out


DATA = {
    "Lab Information": kv("Lab Information", {
        "Name": "آزمایشگاه پتروشیمی تندیس پارس (TPPC)", "LabURL": "https://tppc.ir",
        "Confidence": "95", "LaboratoryAccredited": "1", "Accreditation": "ISO/IEC 17025",
        "AccreditationBody": "NACI", "Physical_Country": "Iran", "Postal_Country": "Iran",
        "Billing_Country": "Iran", "EmailAddress": "", "Physical_City": "", "Postal_City": "",
        "Billing_City": "", "Physical_Address": "", "Postal_Address": "", "Billing_Address": "",
        "Physical_Zip": "", "Postal_Zip": "", "Billing_Zip": ""}),
    "Setup": kv("Setup", {"Currency": "IRR", "DefaultCountry": "IR", "VAT": "9",
                          "MemberDiscount": "0", "ShowPricing": "0", "SamplingWorkflowEnabled": "0"}),
    "Lab Contacts": [{
        "Firstname": fn, "Surname": sn, "Username": u, "Password": PW,
        "Roles": roles, "Groups": groups, "Department_title": dept,
        "JobTitle": job, "EmailAddress": f"{u}@tppc.ir",
    } for fn, sn, u, roles, groups, dept, job in USERS],
}

out = openpyxl.Workbook(); out.remove(out.active)
hf = PatternFill("solid", fgColor="1A3C6E"); hfont = Font(color="FFFFFF", bold=True)
for name in ex.sheetnames:
    keys = [c.value for c in next(ex[name].iter_rows(min_row=1, max_row=1)) if c.value is not None]
    ws = out.create_sheet(name[:31])
    for c, k in enumerate(keys, 1):
        cell = ws.cell(1, c, k); cell.fill = hf; cell.font = hfont
    ws.cell(2, 1, name)
    for c, k in enumerate(keys, 1):
        ws.cell(3, c, k)
    for ri, rec in enumerate(DATA.get(name, []), 4):
        for ci, k in enumerate(keys, 1):
            ws.cell(ri, ci, rec.get(k, ""))
out.save(OUT)
print(f"✅ {OUT} | کاربران: {len(USERS)}")
for fn, sn, u, roles, groups, dept, job in USERS:
    print(f"   {u:9} | {fn} {sn:10} | نقش: {roles}")
