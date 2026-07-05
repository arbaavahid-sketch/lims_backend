"""import افزایشی خانوادهٔ «قیر»: ۵ گرید نفوذی + ۶ سرویس آزمون + حدود (INSO 2950/12505).

ورودی حدها: TPPC_Test_Limits_Master_v2 (استخراج‌شده و اینجا کدگذاری‌شده).
خروجی: TPPC_Bitumen_Import.xlsx  (اسکلت ۶۰ شیت + LabInfo/Setup + SampleTypes + Services + Specs)
افزایشی است؛ روی داده‌های موجود اضافه می‌شود (keywordهای BIT_* یکتا هستند).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

EXAMPLE = "../../senaite.core/src/bika/lims/setupdata/test/test.xlsx"
OUT = "TPPC_Bitumen_Import.xlsx"
LAB = "قیر"
NO_MAX = 999999  # سنتینل «بدون حد بالا» (چون importer، max خالی را ۰ می‌کند)

# سرویس‌های آزمون قیر (keyword یکتا، عنوان، واحد، متد)
SERVICES = [
    ("BIT_PEN",       "درجه نفوذ قیر در ۲۵°C (100g,5s)", "0.1 mm", "INSO 2950 / ASTM D5"),
    ("BIT_FLASH",     "نقطه اشتعال قیر (کلیولند، D92)",  "°C",     "INSO 2954 / ASTM D92"),
    ("BIT_DUCT",      "کشش‌پذیری قیر (داکتیلیتی، D113)",  "cm",     "INSO 3866 / ASTM D113"),
    ("BIT_SOL",       "حلالیت قیر در تری‌کلرواتیلن",      "%",      "INSO 2953 / ASTM D2042"),
    ("BIT_TFOT_PEN",  "نسبت نفوذ باقیمانده پس از TFOT",  "%",      "INSO 2950 / ASTM D5"),
    ("BIT_TFOT_DUCT", "کشش‌پذیری پسماند پس از TFOT",     "cm",     "INSO 3866 / ASTM D113"),
]
SVC_TITLE = {kw: t for kw, t, u, m in SERVICES}

# گریدها: (عنوان نوع‌نمونه، پیشوند، {keyword: (min, max)})  — max None یعنی «بدون حد بالا»
GRADES = [
    ("قیر خالص درجه نفوذ 40-50", "BIT40", {
        "BIT_PEN": (40, 50), "BIT_FLASH": (232, None), "BIT_DUCT": (100, None),
        "BIT_SOL": (99, None), "BIT_TFOT_PEN": (58, None), "BIT_TFOT_DUCT": (100, None)}),
    ("قیر خالص درجه نفوذ 60-70", "BIT60", {
        "BIT_PEN": (60, 70), "BIT_FLASH": (232, None), "BIT_DUCT": (100, None),
        "BIT_SOL": (99, None), "BIT_TFOT_PEN": (54, None), "BIT_TFOT_DUCT": (100, None)}),
    ("قیر خالص درجه نفوذ 85-100", "BIT85", {
        "BIT_PEN": (85, 100), "BIT_FLASH": (232, None), "BIT_DUCT": (100, None),
        "BIT_SOL": (99, None), "BIT_TFOT_PEN": (50, None), "BIT_TFOT_DUCT": (75, None)}),
    ("قیر خالص درجه نفوذ 120-150", "BIT120", {
        "BIT_PEN": (120, 150), "BIT_FLASH": (218, None), "BIT_DUCT": (100, None),
        "BIT_SOL": (99, None), "BIT_TFOT_PEN": (46, None), "BIT_TFOT_DUCT": (50, None)}),
    ("قیر خالص درجه نفوذ 200-300", "BIT200", {
        "BIT_PEN": (200, 300), "BIT_FLASH": (177, None), "BIT_DUCT": (100, None),
        "BIT_SOL": (99, None), "BIT_TFOT_PEN": (40, None)}),
]

ex = openpyxl.load_workbook(EXAMPLE, read_only=True, data_only=True)


def kv(sheet, overrides):
    ws = ex[sheet]
    keys = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value]
    out = []
    for row in list(ws.iter_rows(values_only=True))[3:]:
        rec = dict(zip(keys, row))
        if not rec.get("Field"):
            continue
        if rec["Field"] in overrides:
            rec["Value"] = overrides[rec["Field"]]
        out.append({k: ("" if rec.get(k) is None else rec.get(k)) for k in keys})
    return out


DATA = {}
DATA["Lab Information"] = kv("Lab Information", {
    "Name": "آزمایشگاه پتروشیمی تندیس پارس (TPPC)", "LabURL": "https://tppc.ir",
    "Confidence": "95", "LaboratoryAccredited": "1", "Accreditation": "ISO/IEC 17025",
    "AccreditationBody": "NACI", "Physical_Country": "Iran", "Postal_Country": "Iran",
    "Billing_Country": "Iran", "EmailAddress": "", "Physical_City": "", "Postal_City": "",
    "Billing_City": "", "Physical_Address": "", "Postal_Address": "", "Billing_Address": "",
    "Physical_Zip": "", "Postal_Zip": "", "Billing_Zip": "",
})
DATA["Setup"] = kv("Setup", {"Currency": "IRR", "DefaultCountry": "IR", "VAT": "9",
                             "MemberDiscount": "0", "ShowPricing": "0", "SamplingWorkflowEnabled": "0"})

# نوع‌نمونه‌های گرید قیر
DATA["Sample Types"] = [
    {"title": title, "Prefix": pref, "MinimumVolume": "0 mL", "description": "", "Hazardous": "0"}
    for title, pref, _ in GRADES
]

# سرویس‌های آزمون قیر (در دپارتمان/دستهٔ «قیر»)
DATA["Analysis Services"] = [
    {"title": t, "Keyword": kw, "PointOfCapture": "lab", "AnalysisCategory_title": LAB,
     "Department_title": LAB, "Unit": u, "ManualEntryOfResults": "1", "Accredited": "1", "Price": "0"}
    for kw, t, u, m in SERVICES
]

# مشخصات هر گرید
specs = []
for title, pref, ranges in GRADES:
    for kw, (mn, mx) in ranges.items():
        specs.append({
            "Title": title, "Client_title": "", "SampleType_title": title,
            "service": SVC_TITLE[kw],
            "min": mn if mn is not None else "",
            "max": mx if mx is not None else NO_MAX,  # بدون حد بالا → سنتینل
            "error": "",
        })
DATA["Analysis Specifications"] = specs

# ---------- نوشتن اسکلت کامل ----------
out = openpyxl.Workbook(); out.remove(out.active)
hf = PatternFill("solid", fgColor="1A3C6E"); hfont = Font(color="FFFFFF", bold=True)
for name in ex.sheetnames:
    keys = [c.value for c in next(ex[name].iter_rows(min_row=1, max_row=1)) if c.value is not None]
    ws = out.create_sheet(name[:31])
    for c, k in enumerate(keys, 1):
        cell = ws.cell(1, c, k); cell.fill = hf; cell.font = hfont
    ws.cell(2, 1, name)
    for c, k in enumerate(keys, 1):
        ws.cell(3, c, k)
    for ri, rec in enumerate(DATA.get(name, []), 4):
        for ci, k in enumerate(keys, 1):
            ws.cell(ri, ci, rec.get(k, ""))
out.save(OUT)
print(f"✅ {OUT}")
print(f"   نوع‌نمونه (گرید): {len(DATA['Sample Types'])}")
print(f"   سرویس آزمون قیر: {len(DATA['Analysis Services'])}")
print(f"   ردیف مشخصات: {len(specs)}")
