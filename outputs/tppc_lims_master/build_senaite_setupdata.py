"""تبدیل دادهٔ مادر TPPC به فایل import رسمی SENAITE («Load Setup Data»).

فرمت SENAITE (مهم):
  ردیف ۱ = کلیدهای ماشینی ستون‌ها (importer از این کلید می‌سازد)
  ردیف ۲ و ۳ = بنر/برچسب (توسط get_rows رد می‌شوند)
  ردیف ۴ به بعد = داده
همهٔ شیت‌های استاندارد باید وجود داشته باشند (حتی خالی) وگرنه importer با KeyError می‌افتد.
کلیدهای دقیق هر شیت از فایل نمونهٔ خود SENAITE خوانده می‌شود تا ۱۰۰٪ منطبق باشد.

ورودی‌ها: TPPC_Instruments_final.xlsx، TPPC_AnalysisServices_final.xlsx،
          TPPC_LIMS_Master_Data_units_filled.xlsx
مرجع فرمت: senaite.core/src/bika/lims/setupdata/test/test.xlsx
خروجی:    TPPC_SENAITE_Setup.xlsx
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill

INSTR = "TPPC_Instruments_final.xlsx"
SERV = "TPPC_AnalysisServices_final.xlsx"
MASTER = "TPPC_LIMS_Master_Data_units_filled.xlsx"
EXAMPLE = "../../senaite.core/src/bika/lims/setupdata/test/test.xlsx"
OUT = "TPPC_SENAITE_Setup.xlsx"

# ۶ آزمایشگاه واقعی TPPC (از tppc.ir) — هم دپارتمان، هم دستهٔ آنالیز
LABS = ["هیدروکربن", "روغن", "نفت خام", "ضدیخ", "قیر", "گریس"]


def _norm(s):
    return s.replace("ي", "ی").replace("ى", "ی").replace("ك", "ک")


def LAB(title):
    """نگاشت نام آزمون فارسی به آزمایشگاه واقعی (اولویت: تخصصی‌ها اول).

    آزمون‌های مشترک (نقطه اشتعال، گرانروی…) به آزمایشگاه پیش‌فرض «هیدروکربن»
    می‌روند؛ این نگاشت خودکار و قابل‌بازبینی است.
    """
    t = _norm(title)
    if any(k in t for k in ["ضد یخ", "ضدیخ", "نقطه انجماد", "گلایکول", "گلیکول"]):
        return "ضدیخ"
    if any(k in t for k in ["قیر", "داکتیل", "کشش", "نفوذ", "نرمی", "اسپات", "لکه مواد قیری"]):
        return "قیر"
    if "گریس" in t:
        return "گریس"
    if any(k in t for k in ["نفت خام", "رمزباتوم", "باقیمانده کربن", "کنرادسون",
                            "فشار بخار", "آب و رسوب", "نمک", "کربن باقی"]):
        return "نفت خام"
    # آنالیز عنصری ICP (فلزات/افزودنی‌ها) و عدد اسیدی/قلیایی → عمدتاً آزمایشگاه روغن
    if any(k in t for k in ["طیف سنجی نشر", "طیف سنج نشر", "پلاسما", "نشر اتمی",
                            "عنصری", "عنصر"]):
        return "روغن"
    if any(k in t for k in ["روغن", "روان کننده", "روانکار", "عدد قلیای", "قلیائ",
                            "قلیایی", "قلیاییت", "عدد اسید", "اسیدیته", "خنثی",
                            "جدا شدن آب", "جداپذیری", "افزودنی", "شاخص گرانروی", "کف "]):
        return "روغن"
    return "هیدروکربن"


def load(path, sheet, header_key):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == header_key)
    headers = [str(h).strip() if h else "" for h in rows[hi]]
    return [dict(zip(headers, r)) for r in rows[hi+1:] if any(v not in (None, "") for v in r)]


def s(r, k):
    v = r.get(k)
    return "" if v is None else (v if hasattr(v, "strftime") else str(v).strip())


instruments = load(INSTR, "Instruments", "Instrument Code")
services = load(SERV, "Analysis Services", "Analysis Keyword")
sample_types = load(MASTER, "Sample Types", "Sample Type Code")


def inst_title(r):
    en = s(r, "Name EN") or s(r, "Name FA") or s(r, "Instrument Code")
    return f"{en} ({s(r,'Instrument Code')})"

code2title = {s(r, "Instrument Code"): inst_title(r) for r in instruments}

# ---------- ساخت رکوردهای هر شیت (dict با کلیدهای دقیق SENAITE) ----------
DATA = {}

DATA["Lab Departments"] = [
    {"title": d, "description": "", "LabContact_Username": ""} for d in LABS
]

DATA["Instrument Types"] = [
    {"title": t, "description": ""}
    for t in sorted({s(r, "Instrument Type") for r in instruments if s(r, "Instrument Type")})
]

DATA["Manufacturers"] = [
    {"title": b, "description": ""}
    for b in sorted({s(r, "Manufacturer") for r in instruments if s(r, "Manufacturer")})
]

DATA["Instruments"] = [{
    "title": inst_title(r), "description": s(r, "Name FA"),
    "Type": s(r, "Instrument Type"), "Brand": s(r, "Manufacturer"),
    "Model": s(r, "Model"), "SerialNo": s(r, "Serial Number"),
    "Location": s(r, "Location"),
    "CalibrationExpiryDate": s(r, "Next Calibration Due"),
} for r in instruments]

# دستهٔ آنالیز = همان ۶ آزمایشگاه (هر دسته زیر دپارتمان هم‌نام خودش)
DATA["Analysis Categories"] = [
    {"title": lab, "description": "", "Department_title": lab} for lab in LABS
]

serv_records = []
lab_counts = {}
for r in services:
    code = s(r, "Instrument Code")
    lab = LAB(s(r, "Analysis Title FA"))
    lab_counts[lab] = lab_counts.get(lab, 0) + 1
    serv_records.append({
        "title": s(r, "Analysis Title FA"), "Keyword": s(r, "Analysis Keyword"),
        "PointOfCapture": "lab", "AnalysisCategory_title": lab, "Department_title": lab,
        "Unit": s(r, "Unit"), "ManualEntryOfResults": "1",
        "DefaultInstrument_title": code2title.get(code, ""),
        "MaxTimeAllowed_days": s(r, "Turnaround Days"), "Accredited": "1", "Price": "0",
    })
DATA["Analysis Services"] = serv_records

st_records = []
used = set()
for i, r in enumerate(sample_types, 1):
    status = s(r, "Status").lower()
    if status and status != "active":
        continue
    title = s(r, "Sample Type / Product") or s(r, "Sample Type Code")
    if not title:
        continue
    pref = re.sub(r"[^A-Za-z0-9]", "", s(r, "Sample Type Code")).upper()[:4]
    if not pref or pref in used:
        pref = f"ST{i:03d}"
    used.add(pref)
    st_records.append({"title": title, "Prefix": pref, "MinimumVolume": "0 mL",
                       "description": "", "Hazardous": "0"})
DATA["Sample Types"] = st_records

# ---------- شیت‌های کلید-مقداری الزامی (Lab Information, Setup) ----------
# مقادیر را از فایل نمونه برمی‌داریم و فقط موارد تندیس را override می‌کنیم
# تا هیچ فیلدی که importer می‌خواند جا نماند.
ex = openpyxl.load_workbook(EXAMPLE, read_only=True, data_only=True)


def kv_from_example(sheet, overrides):
    ws = ex[sheet]
    keys = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value]
    records = []
    for row in list(ws.iter_rows(values_only=True))[3:]:  # داده از ردیف ۴
        rec = dict(zip(keys, row))
        field = rec.get("Field")
        if not field:
            continue
        if field in overrides:
            rec["Value"] = overrides[field]
        records.append({k: ("" if rec.get(k) is None else rec.get(k)) for k in keys})
    return records


DATA["Lab Information"] = kv_from_example("Lab Information", {
    "Name": "آزمایشگاه پتروشیمی تندیس پارس (TPPC)",
    "LabURL": "https://tppc.ir",
    "Confidence": "95",
    "LaboratoryAccredited": "1",
    "AccreditationBodyLong": "مرکز ملی تأیید صلاحیت ایران",
    "AccreditationBody": "NACI",
    "AccreditationBodyURL": "naci.gov.ir",
    "Accreditation": "ISO/IEC 17025",
    "AccreditationReference": "",
    "EmailAddress": "",
    "Physical_Address": "", "Physical_City": "", "Physical_State": "",
    "Physical_Zip": "", "Physical_Country": "Iran",
    "Postal_Address": "", "Postal_City": "", "Postal_State": "",
    "Postal_Zip": "", "Postal_Country": "Iran",
    "Billing_Address": "", "Billing_City": "", "Billing_State": "",
    "Billing_Zip": "", "Billing_Country": "Iran",
})

DATA["Setup"] = kv_from_example("Setup", {
    "Currency": "IRR",
    "DefaultCountry": "IR",
    "VAT": "9",
    "MemberDiscount": "0",
    "ShowPricing": "0",
    "SamplingWorkflowEnabled": "0",
    "CategoriseAnalysisServices": "1",
})

# مشخصات گازوئیل را هم بگنجان (تا در re-import کامل از دست نرود)
try:
    gw = openpyxl.load_workbook("TPPC_Specs_Gasoil.xlsx", data_only=True)["Analysis Specifications"]
    grows = list(gw.iter_rows(values_only=True))
    gkeys = [c for c in grows[0] if c]
    specs = []
    for row in grows[3:]:
        rec = dict(zip(gkeys, row))
        if rec.get("SampleType_title") and rec.get("service"):
            specs.append({k: ("" if rec.get(k) is None else rec.get(k)) for k in gkeys})
    DATA["Analysis Specifications"] = specs
except FileNotFoundError:
    pass

out = openpyxl.Workbook()
out.remove(out.active)
hdr_fill = PatternFill("solid", fgColor="1A3C6E")
hdr_font = Font(color="FFFFFF", bold=True)

filled = 0
for name in ex.sheetnames:
    ex_ws = ex[name]
    keys = [c.value for c in next(ex_ws.iter_rows(min_row=1, max_row=1))]
    keys = [k for k in keys if k is not None]
    ws = out.create_sheet(name[:31])
    # ردیف ۱: کلیدهای ماشینی (منطبق با SENAITE)
    for c, k in enumerate(keys, 1):
        cell = ws.cell(1, c, k); cell.fill = hdr_fill; cell.font = hdr_font
    # ردیف ۲: بنر  |  ردیف ۳: تکرار کلید (هر دو توسط importer رد می‌شوند)
    ws.cell(2, 1, name)
    for c, k in enumerate(keys, 1):
        ws.cell(3, c, k)
    # ردیف ۴ به بعد: داده (اگر برای این شیت داریم)
    records = DATA.get(name, [])
    for ri, rec in enumerate(records, 4):
        for ci, k in enumerate(keys, 1):
            v = rec.get(k, "")
            cell = ws.cell(ri, ci, v)
            if hasattr(v, "strftime"):
                cell.number_format = "yyyy-mm-dd"
    if records:
        filled += 1

out.save(OUT)
print(f"✅ ساخته شد: {OUT}  ({len(ex.sheetnames)} شیت، {filled} شیت دارای داده)\n")
for name in ["Lab Departments", "Instrument Types", "Manufacturers", "Instruments",
             "Analysis Categories", "Analysis Services", "Sample Types",
             "Analysis Specifications"]:
    print(f"   {name:24} {len(DATA.get(name, [])):4} ردیف")
print("\n   توزیع ۱۵۸ سرویس بین ۶ آزمایشگاه:")
for lab in LABS:
    print(f"     {lab:12} {lab_counts.get(lab, 0)}")
